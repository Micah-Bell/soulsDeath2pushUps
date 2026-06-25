"""
═══════════════════════════════════════════════════════════════
  SOULS DEATH TRACKER
═══════════════════════════════════════════════════════════════
  Automatically tracks deaths across FromSoftware / Souls games.

  HOW IT WORKS (big picture):
    1. GameDetector  — runs in the background, scanning your
       running processes every few seconds. When it spots a
       known Souls game .exe, it updates the GUI label and
       stores the game name so deaths are tagged correctly.

    2. ScreenshotDetector — also runs in the background on its
       own thread. Every 0.3 s it grabs a screenshot of your
       primary monitor and uses OpenCV template matching to
       compare it against your "you_died.jpg" image. If the
       match score hits the confidence threshold it fires the
       death callback (same one as the F9 hotkey).

    3. ExcelManager — every time a death is recorded it opens
       the .xlsx file, appends a row to the LEFT table (one row
       per individual death) and either creates or updates the
       matching row in the RIGHT table (one row per session,
       running totals). Then saves and closes the file.

    4. DeathTrackerApp (GUI) — the tkinter window that ties
       everything together. It owns the counters, wires up the
       callbacks, and provides manual controls as a fallback.

═══════════════════════════════════════════════════════════════
"""

# ── Standard library ─────────────────────────────────────────
import tkinter as tk                       # built-in GUI toolkit
from tkinter import filedialog, messagebox # file picker + popup dialogs
import os                                  # file path checks
import json                                # saving/loading config
import threading                           # run detectors without freezing the GUI
import time                                # sleep between polls

# ── Third-party ──────────────────────────────────────────────
import openpyxl                            # read/write .xlsx files
from openpyxl.styles import (              # Excel cell styling
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter  # converts col number → letter (e.g. 3 → "C")
from datetime import datetime              # timestamps for each death


# ═══════════════════════════════════════════════════════════════
#  KNOWN GAMES
#  Maps the lowercase .exe filename → human-readable game name.
#  psutil gives us process names, so we match against these keys.
#  Multiple keys can map to the same game (e.g. both "darksouls.exe"
#  and "darksoulsremastered.exe" → "Dark Souls Remastered") because
#  different store versions or launch methods can have different names.
# ═══════════════════════════════════════════════════════════════
KNOWN_GAMES = {
    # ── Elden Ring family ────────────────────────────────────
    "eldenring.exe":               "Elden Ring",
    "nightreign.exe":              "Elden Ring Nightreign",       # standalone spin-off
    "eldenringnightreign.exe":     "Elden Ring Nightreign",       # alternate exe name

    # ── Dark Souls ───────────────────────────────────────────
    "darksouls.exe":               "Dark Souls Remastered",
    "darksoulsr.exe":              "Dark Souls Remastered",
    "darksoulsremastered.exe":     "Dark Souls Remastered",
    "darksoulsii.exe":             "Dark Souls II",
    "darksouls2.exe":              "Dark Souls II",
    "scholar.exe":                 "Dark Souls II: Scholar",      # Scholar of the First Sin edition
    "darksoulsiii.exe":            "Dark Souls III",
    "darksouls3.exe":              "Dark Souls III",

    # ── Other FromSoftware titles ─────────────────────────────
    "sekiro.exe":                  "Sekiro: Shadows Die Twice",
    "armoredcore6.exe":            "Armored Core VI",
    "ac6.exe":                     "Armored Core VI",

    # ── Bloodborne ───────────────────────────────────────────
    # Bloodborne has no PC port so detect common emulators instead.
    "bloodborne.exe":              "Bloodborne",                  # hypothetical native port
    "rpcs3.exe":                   "Bloodborne (RPCS3)",          # PS3 emulator (BB runs here)
    "shadps4.exe":                 "Bloodborne (shadPS4)",        # PS4 emulator

    # ── Demon's Souls ─────────────────────────────────────────
    "demons souls.exe":            "Demon's Souls",
    "demonssouls.exe":             "Demon's Souls",
    "demonssoulsr.exe":            "Demon's Souls Remake",        # PS5 Bluepoint remake
}


# ═══════════════════════════════════════════════════════════════
#  CONFIG
#  Settings are saved to tracker_config.json next to the script
#  so they persist between runs. If the file doesn't exist we
#  fall back to DEFAULT_CFG and create it on first save.
# ═══════════════════════════════════════════════════════════════
CONFIG_FILE = "tracker_config.json"
DEFAULT_CFG = {
    "save_path":       "souls_deaths.xlsx",  # where the Excel file lives
    "template_path":   "you_died.jpg",       # your cropped YOU DIED image
    "detect_interval": 3,     # seconds between game-process scans
    "death_cooldown":  5,     # seconds to wait after a detection before detecting again
                              # (prevents one death screen from triggering multiple counts)
    "confidence":      0.75,  # template match score needed to count as a death (0.0–1.0)
                              # lower = more sensitive but more false positives
}


# ═══════════════════════════════════════════════════════════════
#  COLUMN LAYOUT CONSTANTS
#  openpyxl uses 1-based column numbers, so col 1 = "A", etc.
#  We name them here so the rest of the code is readable instead
#  of being full of magic numbers like ws.cell(row, 7).
#
#  Deaths Log sheet layout:
#    A(1)  B(2)  C(3)  |  D(4)  |  E(5)  F(6)  G(7)  H(8)  I(9)
#    ───────────────────   ─────   ──────────────────────────────────
#    Session Game  Time  │  gap  │  Session  Game  Deaths  Start  End
#    LEFT TABLE (per death)        RIGHT TABLE (per session)
# ═══════════════════════════════════════════════════════════════
LEFT_COLS  = {"session": 1, "game": 2, "time": 3}
RIGHT_COLS = {"session": 5, "game": 6, "deaths": 7, "start": 8, "end": 9}
GAP_COL    = 4          # the dark divider column between the two tables
TS_FMT     = "YYYY-MM-DD HH:MM:SS"   # Excel number format string for timestamps


# ═══════════════════════════════════════════════════════════════
#  EXCEL MANAGER
#  Responsible for everything Excel-related:
#    - Creating the .xlsx file with both sheets on first run
#    - Writing headers and styling them
#    - Appending death rows to the left table
#    - Finding and updating (or creating) session rows in the right table
# ═══════════════════════════════════════════════════════════════
class ExcelManager:

    # ── Colour palette (hex, no '#') ─────────────────────────
    HDR_BG = "1A1A2E"   # dark navy — header row background
    ALT_BG = "13519E"   # slightly lighter navy — alternating row tint
    ACCENT = "C84B31"   # blood red — death count highlight + borders
    TEXT   = "E0E0E038" # grey — regular cell text
    DIM    = "666666"   # mid grey — de-emphasised text (timestamps)
    GAP_BG = "080808"   # near-black — the divider column

    def __init__(self, path):
        self.path = path
        # Only create the file if it doesn't already exist.
        # This means existing data is never overwritten by restarting the tracker.
        if not os.path.exists(path):
            self._create_file()

    # ─────────────────────────────────────────────────────────
    #  FILE CREATION
    # ─────────────────────────────────────────────────────────

    def _create_file(self):
        """
        Build a fresh workbook with two sheets:
          - "Deaths Log"  : the dual-table layout (left = per death, right = per session)
          - "Summary"     : formula-driven grand totals that update automatically
        """
        wb = openpyxl.Workbook()

        # openpyxl creates one sheet called "Sheet" by default; rename it
        ws = wb.active
        ws.title = "Deaths Log"
        self._setup_deaths_sheet(ws)

        # Add a second sheet for the grand-total summary
        summary = wb.create_sheet("Summary")
        self._setup_summary_sheet(summary)

        wb.save(self.path)

    def _setup_deaths_sheet(self, ws):
        """
        Configure column widths, write the header row, and freeze the top row
        so it stays visible when you scroll down through hundreds of deaths.
        """
        # Set each column's width in characters. Column 4 (the gap) is narrow on purpose.
        widths = {1: 10, 2: 24, 3: 22, 4: 3, 5: 10, 6: 24, 7: 10, 8: 22, 9: 22}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # Write both header sections in one call
        self._write_section_header(
            ws, row=1,
            left_labels=["Session", "Game", "Death Time"],
            right_labels=["Session", "Game", "Deaths", "Start Time", "End Time"]
        )

        ws.row_dimensions[1].height = 22  # slightly taller header row

        # freeze_panes = "A2" means row 1 is frozen (the header stays put while scrolling)
        ws.freeze_panes = "A2"

        # Make the gap column header match the gap background
        ws.cell(1, GAP_COL).fill = PatternFill("solid", fgColor=self.GAP_BG)

    def _write_section_header(self, ws, row, left_labels, right_labels):
        """
        Write styled header cells for both the left and right tables.
        The gap column (D) gets a different fill so it looks like a divider.
        """
        hdr_fill   = PatternFill("solid", fgColor=self.HDR_BG)
        hdr_font   = Font(bold=True, color="FFFFFF", size=10)
        # A red bottom border gives the header a visual "separator" feel
        hdr_border = Border(bottom=Side(style="medium", color=self.ACCENT))

        # Write left table headers (columns 1, 2, 3)
        for col_idx, label in zip([1, 2, 3], left_labels):
            c = ws.cell(row, col_idx)
            c.value     = label
            c.fill      = hdr_fill
            c.font      = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = hdr_border

        # The gap column gets only a background, no text
        ws.cell(row, GAP_COL).fill = PatternFill("solid", fgColor=self.GAP_BG)

        # Write right table headers (columns 5, 6, 7, 8, 9)
        for col_idx, label in zip([5, 6, 7, 8, 9], right_labels):
            c = ws.cell(row, col_idx)
            c.value     = label
            c.fill      = hdr_fill
            c.font      = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = hdr_border

    def _setup_summary_sheet(self, ws):
        """
        Build the Summary sheet. Every cell in column B is an Excel formula
        that reads from the Deaths Log sheet — so this sheet updates live
        whenever you open the file after new deaths have been logged.

        We reference two different data ranges:
          - 'Deaths Log'!A:A / C:C  → left table (one row per death)
          - 'Deaths Log'!E:E / G:G  → right table (one row per session)
        """
        dark = PatternFill("solid", fgColor="0D0D0D")

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 22

        # Title merged across A1:B1
        ws["A1"] = "SOULS DEATH TRACKER — SUMMARY"
        ws["A1"].font = Font(bold=True, size=13, color=self.ACCENT)
        ws.merge_cells("A1:B1")
        ws["A1"].fill = dark
        ws["B1"].fill = dark

        # Column subheaders
        ws["A2"] = "Stat"
        ws["B2"] = "Value"
        for cell in [ws["A2"], ws["B2"]]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=self.HDR_BG)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = Border(bottom=Side(style="medium", color=self.ACCENT))

        # Each tuple is (label, Excel formula).
        # The formulas are explained inline below:
        stats = [
            (
                "Total Deaths",
                # COUNTA counts non-empty cells in column A of Deaths Log.
                # We subtract 1 to exclude the header row.
                "=COUNTA('Deaths Log'!A:A)-1"
            ),
            (
                "Total Sessions",
                # The session number in the right table (col E) always increases,
                # so MAX of that column = total number of sessions played.
                "=MAX('Deaths Log'!E:E)"
            ),
            (
                "Most Deaths in Session",
                # Column G of the right table holds per-session death counts.
                # IF guard prevents showing 0 when no data exists yet.
                "=IF(MAX('Deaths Log'!G:G)>0,MAX('Deaths Log'!G:G),0)"
            ),
            (
                "Least Deaths in Session",
                # MINIFS finds the minimum value in G that is >0,
                # so an empty column doesn't return 0 incorrectly.
                "=IF(COUNTA('Deaths Log'!G:G)>1,"
                "MINIFS('Deaths Log'!G2:G9999,'Deaths Log'!G2:G9999,\">0\"),0)"
            ),
            (
                "Avg Deaths / Session",
                # Total deaths (rows in left table) divided by session count.
                # IF guard avoids a #DIV/0! error when no sessions exist yet.
                "=IF(MAX('Deaths Log'!E:E)>0,"
                "(COUNTA('Deaths Log'!A:A)-1)/MAX('Deaths Log'!E:E),0)"
            ),
            (
                "Overall First Death",
                # Column C of the left table holds death timestamps.
                # MIN of a datetime column gives the earliest = first ever death.
                "=IF(COUNTA('Deaths Log'!C:C)>1,MIN('Deaths Log'!C2:C9999),\"\")"
            ),
            (
                "Overall Last Death",
                # MAX of the same column = most recent death timestamp.
                "=IF(COUNTA('Deaths Log'!C:C)>1,MAX('Deaths Log'!C2:C9999),\"\")"
            ),
        ]

        # Sheet rows 3–9 (enumerate starts at 3 to skip title + subheader)
        # Rows that hold timestamps need a special number format so Excel
        # displays them as dates rather than raw serial numbers.
        timestamp_rows = {8, 9}   # "Overall First Death" and "Overall Last Death"

        for i, (label, formula) in enumerate(stats, start=3):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(bold=True, color="AAAAAA", size=10)
            ws[f"A{i}"].fill = dark

            ws[f"B{i}"] = formula
            ws[f"B{i}"].font      = Font(color=self.TEXT, size=10)
            ws[f"B{i}"].fill      = dark
            ws[f"B{i}"].alignment = Alignment(horizontal="center")

            # Apply timestamp display format to date rows
            if i in timestamp_rows:
                ws[f"B{i}"].number_format = TS_FMT

            # Show averages with one decimal place
            if "Avg" in label:
                ws[f"B{i}"].number_format = "0.0"


    # ─────────────────────────────────────────────────────────
    #  DEATH LOGGING  (called every time a death is detected)
    # ─────────────────────────────────────────────────────────

    def log_death(self, game, session, death_time=None):
        """
        The main entry point for recording a death.

        Steps:
          1. Open the workbook from disk (so we always have the latest state)
          2. Determine the timestamp (now, unless one was passed in for testing)
          3. Append a new row to the LEFT table (per-death log)
          4. Find the matching session row in the RIGHT table:
               - If found → increment its death count and update End Time
               - If not found → create a new row for this session
          5. Save the workbook back to disk
        """
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Deaths Log"]
        ts = death_time or datetime.now()

        # ── Step 3: Left table ──────────────────────────────
        # Walk down column A (session col) until we find an empty cell.
        # That's where the new death row goes.
        left_row = self._next_empty_row(ws, col=LEFT_COLS["session"], start=2)
        self._write_left_row(ws, left_row, session, game, ts)

        # ── Step 4: Right table ─────────────────────────────
        # Search column E for a row whose session number matches ours.
        right_row = self._find_session_row(ws, session)

        if right_row is None:
            # This is the first death in this session — create a new row
            right_row = self._next_empty_row(ws, col=RIGHT_COLS["session"], start=2)
            self._write_right_row_new(ws, right_row, session, game, ts)
        else:
            # Session already exists — just bump the count and update End Time
            self._update_right_row(ws, right_row, ts)

        # ── Step 5: Save ────────────────────────────────────
        wb.save(self.path)

    # ─────────────────────────────────────────────────────────
    #  HELPER: ROW SEARCH UTILITIES
    # ─────────────────────────────────────────────────────────

    def _next_empty_row(self, ws, col, start):
        """
        Walk down a column starting at `start` until we hit a cell with no value.
        Returns that row number — this is where we'll write next.

        Example: if rows 2, 3, 4 in col A are filled, returns 5.
        """
        row = start
        while ws.cell(row, col).value is not None:
            row += 1
        return row

    def _find_session_row(self, ws, session):
        """
        Scan the right table's session column (col E) looking for a row
        whose value matches the current session number.

        Returns the row number if found, or None if this session is new.
        This is how we know whether to create a new right-table row
        or update an existing one.
        """
        row = 2  # start below the header
        while ws.cell(row, RIGHT_COLS["session"]).value is not None:
            if ws.cell(row, RIGHT_COLS["session"]).value == session:
                return row   # found it
            row += 1
        return None   # reached an empty cell — session doesn't exist yet


    # ─────────────────────────────────────────────────────────
    #  HELPER: CELL STYLING
    # ─────────────────────────────────────────────────────────

    def _style_left(self, ws, row):
        """
        Apply formatting to a left-table row (the per-death log).
        Even-numbered rows get a slightly lighter background (ALT_BG)
        to create a zebra-stripe effect that makes rows easier to read.
        The gap column is always kept dark so the visual divider is consistent.
        """
        # Zebra stripe: every other row gets a tinted background
        alt = PatternFill("solid", fgColor=self.ALT_BG) if row % 2 == 0 else None

        for col in [1, 2, 3]:
            c = ws.cell(row, col)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(color=self.TEXT, size=10)
            if alt:
                c.fill = alt

        # The death time cell needs a number format so Excel shows it as a
        # human-readable date+time rather than a float (Excel stores dates as
        # floats internally — e.g. 46224.8 = some date in 2026)
        ws.cell(row, 3).number_format = TS_FMT
        ws.row_dimensions[row].height = 18

        # Keep the gap column visually dark on every data row
        ws.cell(row, GAP_COL).fill = PatternFill("solid", fgColor=self.GAP_BG)

    def _style_right(self, ws, row):
        """
        Apply formatting to a right-table row (the per-session summary).
        Death count is highlighted in red (ACCENT) to make it stand out.
        Start/End times are dimmed because they're supporting info.
        """
        alt = PatternFill("solid", fgColor=self.ALT_BG) if row % 2 == 0 else None

        # Apply alignment and optional zebra stripe to all right-table columns
        for col in [5, 6, 7, 8, 9]:
            c = ws.cell(row, col)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if alt:
                c.fill = alt

        # Individual cell fonts
        ws.cell(row, RIGHT_COLS["session"]).font = Font(color=self.TEXT, size=10)
        ws.cell(row, RIGHT_COLS["game"]).font    = Font(color=self.TEXT, size=10)
        ws.cell(row, RIGHT_COLS["deaths"]).font  = Font(bold=True, color=self.ACCENT, size=10)  # red + bold
        ws.cell(row, RIGHT_COLS["start"]).font   = Font(color=self.DIM, size=9)   # dimmed
        ws.cell(row, RIGHT_COLS["end"]).font     = Font(color=self.DIM, size=9)   # dimmed

        # Both time columns need the datetime display format
        ws.cell(row, RIGHT_COLS["start"]).number_format = TS_FMT
        ws.cell(row, RIGHT_COLS["end"]).number_format   = TS_FMT


    # ─────────────────────────────────────────────────────────
    #  HELPER: ROW WRITERS
    # ─────────────────────────────────────────────────────────

    def _write_left_row(self, ws, row, session, game, ts):
        """Write one death entry into the left table and style it."""
        ws.cell(row, LEFT_COLS["session"]).value = session
        ws.cell(row, LEFT_COLS["game"]).value    = game
        ws.cell(row, LEFT_COLS["time"]).value    = ts
        self._style_left(ws, row)

    def _write_right_row_new(self, ws, row, session, game, ts):
        """
        Create a brand new session row in the right table.
        Deaths starts at 1, and Start Time = End Time = this death's timestamp
        (they'll diverge as more deaths are logged in the same session).
        """
        ws.cell(row, RIGHT_COLS["session"]).value = session
        ws.cell(row, RIGHT_COLS["game"]).value    = game
        ws.cell(row, RIGHT_COLS["deaths"]).value  = 1    # first death in this session
        ws.cell(row, RIGHT_COLS["start"]).value   = ts   # session started here
        ws.cell(row, RIGHT_COLS["end"]).value     = ts   # will be updated on future deaths
        self._style_right(ws, row)

    def _update_right_row(self, ws, row, ts):
        """
        Increment the death count for an existing session and push the End Time
        forward to this death's timestamp. Start Time is never changed —
        it stays as the timestamp of the very first death in the session.
        """
        deaths_cell = ws.cell(row, RIGHT_COLS["deaths"])
        # Guard against None just in case the cell is empty for some reason
        deaths_cell.value = (deaths_cell.value or 0) + 1
        ws.cell(row, RIGHT_COLS["end"]).value = ts   # latest death = new end time
        self._style_right(ws, row)   # re-apply style (needed when updating existing cells)


# ═══════════════════════════════════════════════════════════════
#  GAME DETECTOR
#  Polls the system process list on a background thread every
#  `detect_interval` seconds. When it finds a known game exe
#  it fires the on_change callback with the game name.
#  When the game closes, on_change is called with None.
#
#  Why a background thread?
#  If we did this on the main thread it would freeze the GUI
#  for a couple of seconds every poll. threading.Thread with
#  daemon=True means this thread dies automatically when the
#  main program exits — no cleanup needed.
# ═══════════════════════════════════════════════════════════════
class GameDetector:

    def __init__(self, on_change):
        """
        on_change: a callable that receives either a game name string
                   or None (when no game is running).
        """
        self.on_change = on_change
        self.running   = False
        self.current   = None    # tracks what game was last seen so we only
                                 # fire the callback when something changes
        self._thread   = None

    def start(self):
        """Kick off the background polling thread."""
        self.running = True
        # daemon=True: thread exits automatically when the main program closes
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()

    def stop(self):
        """Signal the loop to exit on its next iteration."""
        self.running = False

    def _loop(self):
        """
        The actual polling logic. Runs forever until stop() is called.

        psutil.process_iter(["name"]) is efficient — it only fetches the
        process name attribute rather than all process info. We lowercase
        the name before comparing because Windows process names can vary
        in capitalisation between versions/launchers.
        """
        try:
            import psutil
        except ImportError:
            # If psutil isn't installed the game detector simply does nothing.
            # The rest of the app still works (manual F9, screenshot detection).
            return

        while self.running:
            found = None

            # Iterate over every running process
            for proc in psutil.process_iter(["name"]):
                name = (proc.info["name"] or "").lower()
                if name in KNOWN_GAMES:
                    found = KNOWN_GAMES[name]
                    break   # stop scanning once we've found a match

            # Only fire the callback if the game status changed
            # (avoids spamming the GUI with redundant updates every 3 seconds)
            if found != self.current:
                self.current = found
                self.on_change(found)

            time.sleep(DEFAULT_CFG["detect_interval"])


# ═══════════════════════════════════════════════════════════════
#  SCREENSHOT DETECTOR
#  Uses OpenCV template matching to find the "YOU DIED" screen.
#
#  HOW TEMPLATE MATCHING WORKS:
#    - You provide a small cropped image of the YOU DIED text
#      (your "template")
#    - Every 0.3 s we grab a screenshot of your primary monitor
#    - OpenCV slides the template across the screenshot pixel by pixel
#      and measures how similar each position is (using normalised
#      cross-correlation, TM_CCOEFF_NORMED)
#    - The result is a score from 0.0 (no match) to 1.0 (perfect match)
#    - If that score hits our confidence threshold, a death is recorded
#
#  COOLDOWN:
#    After a match we wait `cooldown` seconds before checking again.
#    Without this, a single 5-second death screen would trigger 16
#    deaths (one every 0.3 s). The cooldown ensures each screen = 1 death.
# ═══════════════════════════════════════════════════════════════
class ScreenshotDetector:

    def __init__(self, on_death, template_path, confidence=0.75, cooldown=5):
        """
        on_death:      callback with no arguments — called when a death is detected
        template_path: path to the YOU DIED image file
        confidence:    match threshold (0.0–1.0). 0.75 = 75% similar
        cooldown:      seconds to ignore the screen after a match
        """
        self.on_death      = on_death
        self.template_path = template_path
        self.confidence    = confidence
        self.cooldown      = cooldown
        self.running       = False
        self._thread       = None

    def start(self):
        """
        Pre-flight checks before starting the detection thread:
          1. Does the template image file exist?
          2. Are the required libraries (cv2, numpy, mss) installed?
        Returns True if we successfully started, False otherwise.
        """
        if not os.path.exists(self.template_path):
            return False   # caller handles the error message

        try:
            import cv2, numpy, mss  # noqa — just checking they're importable
        except ImportError:
            return False   # caller will show "install these packages" message

        self.running = True
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        return True

    def stop(self):
        """Signal the detection loop to exit."""
        self.running = False

    def _loop(self):
        """
        The screenshot detection loop. Runs on a background thread.

        Flow each iteration:
          1. Check if we're still in cooldown → skip if so
          2. Grab a screenshot of the primary monitor (index 1 in mss)
          3. Convert to greyscale — colour info is irrelevant and greyscale
             is faster to process
          4. Run template matching
          5. Check the max score against our threshold
          6. If matched: update last_hit timestamp and fire the callback
        """
        import cv2
        import numpy as np
        import mss

        # Load the template once at startup rather than every frame
        # IMREAD_GRAYSCALE loads it as greyscale directly, matching our screenshot conversion
        template = cv2.imread(self.template_path, cv2.IMREAD_GRAYSCALE)
        if template is None:
            return   # file exists but couldn't be decoded as an image

        last_hit = 0       # unix timestamp of last confirmed death
        sct = mss.mss()    # mss screen capture context

        while self.running:
            now = time.time()

            # ── Cooldown check ──────────────────────────────
            # If we just detected a death, sleep briefly and loop back
            # rather than scanning the screen (it's still showing YOU DIED)
            if now - last_hit < self.cooldown:
                time.sleep(0.2)
                continue

            # ── Capture screenshot ──────────────────────────
            try:
                # monitors[1] is the primary monitor. monitors[0] is a
                # virtual "all monitors combined" bounding box.
                screen = np.array(sct.grab(sct.monitors[1]))
            except Exception:
                # mss can occasionally fail (e.g. if the monitor config changes)
                time.sleep(1)
                continue

            # ── Pre-process ─────────────────────────────────
            # Convert from BGRA (what mss captures) to greyscale
            gray = cv2.cvtColor(screen, cv2.COLOR_BGRA2GRAY)

            # ── Template matching ───────────────────────────
            # TM_CCOEFF_NORMED: normalised cross-correlation
            # Returns a 2D array of match scores for every possible position
            result = cv2.matchTemplate(gray, template, cv2.TM_CCOEFF_NORMED)

            # We only care about the maximum score anywhere on the screen
            # minMaxLoc returns (min_val, max_val, min_loc, max_loc)
            _, max_val, _, _ = cv2.minMaxLoc(result)

            # ── Threshold check ─────────────────────────────
            if max_val >= self.confidence:
                last_hit = time.time()   # start the cooldown timer
                self.on_death()          # fire the death callback

            # Pause briefly between frames to avoid maxing out the CPU.
            # 0.3 s = ~3 checks per second, which is plenty for a death screen
            # that usually stays up for 3-5 seconds.
            time.sleep(0.3)


# ═══════════════════════════════════════════════════════════════
#  DEATH TRACKER APP  (the GUI)
#  Built with tkinter — Python's built-in GUI library.
#
#  tkinter uses a "widget tree" model:
#    root window → frames → labels/buttons/etc.
#  Each widget is packed into its parent using .pack() or .grid().
#
#  tk.IntVar / tk.StringVar are special variables that automatically
#  update any Label/Spinbox that's linked to them via textvariable=.
#  So when we do self.total_deaths.set(5), any Label showing it
#  instantly updates to "5" without us doing anything else.
# ═══════════════════════════════════════════════════════════════
class DeathTrackerApp:

    # ── UI colour constants ──────────────────────────────────
    BG     = "#0D0D0D"   # window background (near-black)
    PANEL  = "#1A1A2E"   # card/panel background (dark navy)
    ACCENT = "#C84B31"   # blood red highlights
    TEXT   = "#E0E0E0"   # main text colour
    DIM    = "#777777"   # secondary / de-emphasised text
    GREEN  = "#4CAF50"   # used when game is detected successfully

    def __init__(self, root):
        """
        Set up the app: load config, create state variables,
        build the UI, then start the background detectors.
        """
        self.root = root
        self.root.title("☠ Souls Death Tracker")
        self.root.configure(bg=self.BG)
        self.root.resizable(False, False)   # fixed window size

        # Load saved settings (or defaults if first run)
        self.cfg = self._load_cfg()

        # ── Reactive state variables ─────────────────────────
        # tk.IntVar / tk.StringVar: any widget linked via textvariable=
        # will automatically display the current value whenever it changes.
        self.total_deaths = tk.IntVar(value=0)
        self.session_num  = tk.IntVar(value=1)
        self.game_var     = tk.StringVar(value="No game detected")
        self.status_var   = tk.StringVar(value="Starting up...")
        self.tmpl_var     = tk.StringVar(value=self.cfg["template_path"])

        # ── Core components ──────────────────────────────────
        self.excel    = ExcelManager(self.cfg["save_path"])
        self.detector = None   # ScreenshotDetector instance (created later)
        self.game_det = GameDetector(on_change=self._on_game_change)

        # ── Startup sequence ─────────────────────────────────
        self._build_ui()           # construct all widgets
        self._bind_hotkeys()       # register F9 / F10
        self.game_det.start()      # begin scanning for game processes
        self._try_start_screenshot()  # start screenshot detector if template exists
        self.status_var.set("Ready — watching for game & deaths")

    # ─────────────────────────────────────────────────────────
    #  CONFIG  LOAD / SAVE
    # ─────────────────────────────────────────────────────────

    def _load_cfg(self):
        """
        Load config from the JSON file if it exists.
        The {**DEFAULT_CFG, **loaded} pattern merges them so that
        any new keys added to DEFAULT_CFG in a future version of
        the script are still present even if the saved file is old.
        """
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE) as f:
                return {**DEFAULT_CFG, **json.load(f)}
        return DEFAULT_CFG.copy()

    def _save_cfg(self):
        """Persist current settings so they survive restarts."""
        self.cfg["template_path"] = self.tmpl_var.get()
        with open(CONFIG_FILE, "w") as f:
            json.dump(self.cfg, f, indent=2)


    # ─────────────────────────────────────────────────────────
    #  UI CONSTRUCTION
    # ─────────────────────────────────────────────────────────

    def _build_ui(self):
        """
        Build every widget in the window top-to-bottom using .pack().
        pack() stacks widgets vertically by default; side="left" stacks
        horizontally within a row frame.
        """
        pad = {"padx": 14, "pady": 5}   # reusable padding shorthand

        # ── App title ────────────────────────────────────────
        tk.Label(self.root, text="☠  SOULS DEATH TRACKER",
                 bg=self.BG, fg=self.ACCENT,
                 font=("Segoe UI", 15, "bold")).pack(pady=(16, 2))

        # ── Detected game panel ──────────────────────────────
        # This panel shows which game is currently running.
        # game_label's text is bound to self.game_var — it updates
        # automatically when the GameDetector fires its callback.
        game_frame = tk.Frame(self.root, bg=self.PANEL,
                              highlightbackground="#333", highlightthickness=1)
        game_frame.pack(fill="x", padx=14, pady=(4, 8))
        tk.Label(game_frame, text="DETECTED GAME", bg=self.PANEL, fg=self.DIM,
                 font=("Segoe UI", 7, "bold")).pack(pady=(6, 0))
        self.game_label = tk.Label(game_frame, textvariable=self.game_var,
                                   bg=self.PANEL, fg=self.TEXT,
                                   font=("Segoe UI", 12, "bold"))
        self.game_label.pack(pady=(0, 6))

        # ── Death counter panel ───────────────────────────────
        # The big number. Bound to self.total_deaths so it updates instantly.
        counter_frame = tk.Frame(self.root, bg=self.PANEL,
                                 highlightbackground=self.ACCENT, highlightthickness=1)
        counter_frame.pack(fill="x", padx=14, pady=4)
        tk.Label(counter_frame, text="TOTAL DEATHS", bg=self.PANEL, fg=self.DIM,
                 font=("Segoe UI", 8, "bold")).pack(pady=(10, 0))
        tk.Label(counter_frame, textvariable=self.total_deaths,
                 bg=self.PANEL, fg=self.ACCENT,
                 font=("Segoe UI", 64, "bold")).pack()
        tk.Label(counter_frame, text="(per-session totals tracked in Excel)",
                 bg=self.PANEL, fg=self.DIM,
                 font=("Segoe UI", 7)).pack(pady=(0, 8))

        # ── Session row ───────────────────────────────────────
        # Spinbox lets the user manually set the session number if needed.
        # Normally you'd just press "New Session" which auto-increments it.
        sess_frame = tk.Frame(self.root, bg=self.BG)
        sess_frame.pack(fill="x", **pad)
        tk.Label(sess_frame, text="Session:", bg=self.BG, fg=self.DIM,
                 font=("Segoe UI", 9)).pack(side="left")
        tk.Spinbox(sess_frame, from_=1, to=999, textvariable=self.session_num,
                   width=5, bg=self.PANEL, fg=self.TEXT,
                   buttonbackground=self.PANEL, relief="flat").pack(side="left", padx=6)
        self._btn(sess_frame, "New Session", self.PANEL, self._new_session,
                  fg=self.DIM).pack(side="left", padx=4)

        # ── Action buttons ────────────────────────────────────
        btn_frame = tk.Frame(self.root, bg=self.BG)
        btn_frame.pack(pady=8)
        self._btn(btn_frame, "+DEATH  [F9]", self.ACCENT, self.record_death).pack(side="left", padx=4)
        self._btn(btn_frame, "UNDO  [F10]",  self.PANEL,  self.undo_death, fg=self.DIM).pack(side="left", padx=4)

        # ── Template image picker ─────────────────────────────
        # Clicking the label opens a file picker so you can choose your
        # you_died.jpg. tmpl_status shows ✔ or ✗ depending on whether
        # the file currently exists on disk.
        tmpl_frame = tk.Frame(self.root, bg=self.BG)
        tmpl_frame.pack(fill="x", padx=14, pady=(4, 2))
        tk.Label(tmpl_frame, text="Death image:", bg=self.BG, fg=self.DIM,
                 font=("Segoe UI", 9)).pack(side="left")
        self.tmpl_label = tk.Label(tmpl_frame, textvariable=self.tmpl_var,
                                   bg=self.BG, fg=self.DIM,
                                   font=("Segoe UI", 8), cursor="hand2")
        self.tmpl_label.pack(side="left", padx=6)
        self.tmpl_label.bind("<Button-1>", lambda _: self._pick_template())   # left click
        self.tmpl_status = tk.Label(tmpl_frame, text="", bg=self.BG, font=("Segoe UI", 9))
        self.tmpl_status.pack(side="left")
        self._refresh_tmpl_status()   # show ✔ or ✗ immediately on startup

        # ── Excel path picker ─────────────────────────────────
        # Clicking this opens a save-as dialog to redirect the Excel file.
        path_frame = tk.Frame(self.root, bg=self.BG)
        path_frame.pack(fill="x", padx=14, pady=(2, 4))
        tk.Label(path_frame, text="Excel:", bg=self.BG, fg=self.DIM,
                 font=("Segoe UI", 9)).pack(side="left")
        self.path_label = tk.Label(path_frame, text=self.cfg["save_path"],
                                   bg=self.BG, fg=self.DIM,
                                   font=("Segoe UI", 8), cursor="hand2")
        self.path_label.pack(side="left", padx=6)
        self.path_label.bind("<Button-1>", lambda _: self._pick_excel())

        # ── Status bar ────────────────────────────────────────
        # One-line info strip at the bottom. Shows last death time,
        # errors, mode changes, etc.
        tk.Label(self.root, textvariable=self.status_var,
                 bg=self.BG, fg=self.DIM,
                 font=("Segoe UI", 8)).pack(pady=(4, 14))

    def _btn(self, parent, text, bg, cmd, fg=None):
        """
        Factory helper that creates a styled flat button.
        Centralising button creation means changing the style
        in one place affects all buttons.
        """
        return tk.Button(parent, text=text, command=cmd,
                         bg=bg, fg=fg or self.TEXT,
                         relief="flat",      # no 3D border
                         padx=12, pady=7,
                         font=("Segoe UI", 9, "bold"),
                         cursor="hand2",     # pointer cursor on hover
                         activebackground=self.ACCENT,   # turns red when clicked
                         activeforeground="white")


    # ─────────────────────────────────────────────────────────
    #  HOTKEYS
    # ─────────────────────────────────────────────────────────

    def _bind_hotkeys(self):
        """
        Bind F9 and F10 to the root window.
        The lambda swallows the event object (tkinter always passes one)
        since our methods don't need it.
        Note: these hotkeys only fire when the tracker window is focused.
        For global hotkeys (work in-game) you'd need the `keyboard` library.
        """
        self.root.bind("<F9>",  lambda _: self.record_death())
        self.root.bind("<F10>", lambda _: self.undo_death())


    # ─────────────────────────────────────────────────────────
    #  CALLBACKS
    # ─────────────────────────────────────────────────────────

    def _on_game_change(self, game):
        """
        Called by GameDetector (on a background thread) whenever the detected
        game changes. We use root.after(0, ...) to safely schedule the UI
        update on the main thread — tkinter is NOT thread-safe, so we can't
        update widgets directly from a background thread.

        root.after(0, fn) queues fn to run on the next event loop tick.
        """
        def update():
            if game:
                self.game_var.set(game)
                self.game_label.config(fg=self.GREEN)    # green = game found
                self.status_var.set(f"✔ {game} detected")
            else:
                self.game_var.set("No game detected")
                self.game_label.config(fg=self.DIM)      # grey = no game
                self.status_var.set("Waiting for a Souls game to launch...")
        self.root.after(0, update)

    def record_death(self):
        """
        The core action — called by F9, the +DEATH button, or the
        ScreenshotDetector callback.

        Order of operations:
          1. Increment the in-memory counter (updates the big number instantly)
          2. Log the death to Excel (slightly slower — involves file I/O)
          3. Flash the window border red as visual feedback
        """
        self.total_deaths.set(self.total_deaths.get() + 1)
        total   = self.total_deaths.get()
        game    = self.game_var.get()
        session = self.session_num.get()

        try:
            self.excel.log_death(game=game, session=session)
            self.status_var.set(f"☠  Death #{total} logged  —  {datetime.now().strftime('%H:%M:%S')}")
        except Exception as e:
            # Don't crash the app if Excel fails (e.g. file is open in Excel)
            self.status_var.set(f"⚠ Excel error: {e}")

        self._flash()

    def undo_death(self):
        """
        Decrement the GUI counter. We can't undo the Excel row automatically
        because we'd have to know which row to delete — and with the dual-table
        layout, we'd also need to update the session totals. Easier to just
        tell the user to remove it manually.
        """
        if self.total_deaths.get() <= 0:
            return   # nothing to undo
        self.total_deaths.set(self.total_deaths.get() - 1)
        self.status_var.set("↩ GUI count undone — remove the Excel row manually if needed")

    def _new_session(self):
        """
        Increment the session number. The ExcelManager will automatically
        create a new row in the right table the next time a death is logged
        with the new session number.
        """
        self.session_num.set(self.session_num.get() + 1)
        self.status_var.set(f"Session {self.session_num.get()} started")

    def _flash(self):
        """
        Briefly highlight the window border in red when a death is recorded.
        root.after(250, ...) schedules the border removal 250 ms later,
        creating a quick flash effect without blocking.
        """
        self.root.configure(highlightbackground=self.ACCENT, highlightthickness=2)
        self.root.after(250, lambda: self.root.configure(highlightthickness=0))


    # ─────────────────────────────────────────────────────────
    #  TEMPLATE / SCREENSHOT MANAGEMENT
    # ─────────────────────────────────────────────────────────

    def _pick_template(self):
        """
        Open a file picker to let the user choose their YOU DIED image.
        After picking, save it to config, update the status indicator,
        and restart the screenshot detector with the new template.
        """
        path = filedialog.askopenfilename(
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp"), ("All files", "*.*")],
        )
        if path:
            self.tmpl_var.set(path)        # updates the displayed path label
            self.cfg["template_path"] = path
            self._save_cfg()
            self._refresh_tmpl_status()    # update the ✔ / ✗ indicator
            self._try_start_screenshot()   # restart detector with new template

    def _refresh_tmpl_status(self):
        """
        Check if the template file actually exists and show ✔ or ✗ accordingly.
        Called on startup and whenever the template path changes.
        """
        if os.path.exists(self.tmpl_var.get()):
            self.tmpl_status.config(text="✔", fg=self.GREEN)
        else:
            self.tmpl_status.config(text="✗ not found — click to set", fg="#FF6B6B")

    def _try_start_screenshot(self):
        """
        Stop any existing detector, then start a new one if the template file exists.
        This is called on startup and whenever the template path changes.

        If the required libraries (mss, opencv) aren't installed, start()
        returns False and we show a helpful error in the status bar.
        """
        # Stop the old detector if one is running
        if self.detector:
            self.detector.stop()
            self.detector = None

        path = self.tmpl_var.get()
        if not os.path.exists(path):
            return   # no template = can't start, but that's okay (manual mode still works)

        self.detector = ScreenshotDetector(
            on_death=self.record_death,        # callback: same as pressing F9
            template_path=path,
            confidence=self.cfg["confidence"],
            cooldown=self.cfg["death_cooldown"],
        )

        if self.detector.start():
            self.status_var.set("👁 Auto-detection active")
        else:
            self.status_var.set("⚠ Install mss + opencv-python for auto-detection")
            self.detector = None


    # ─────────────────────────────────────────────────────────
    #  EXCEL PATH
    # ─────────────────────────────────────────────────────────

    def _pick_excel(self):
        """
        Let the user choose a different save location for the Excel file.
        Creates a fresh ExcelManager pointing at the new path (which will
        create the file with headers if it doesn't exist yet).
        """
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.cfg["save_path"] = path
            self.excel = ExcelManager(path)   # new manager = new (or existing) file
            self.path_label.config(text=path)
            self._save_cfg()


# ═══════════════════════════════════════════════════════════════
#  ENTRY POINT
#  This block only runs when you execute the file directly
#  (not when it's imported as a module). Standard Python pattern.
# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("400x440")   # width x height in pixels

    # ── Dark title bar (Windows only) ────────────────────────
    # On Windows 10/11 we can call a DWM (Desktop Window Manager) API
    # to make the title bar dark, matching the rest of the app.
    # We wrap this in a try/except because:
    #   a) It only exists on Windows (not Mac/Linux)
    #   b) The attribute ID (20) only works on Windows 10 build 19041+
    # If it fails for any reason, we just proceed with the default title bar.
    try:
        from ctypes import windll, c_int, sizeof, byref
        val = c_int(1)   # 1 = enable dark mode
        windll.dwmapi.DwmSetWindowAttribute(
            windll.user32.GetForegroundWindow(),   # handle to our window
            20,            # DWMWA_USE_IMMERSIVE_DARK_MODE attribute
            byref(val),    # pointer to our value
            sizeof(val)    # size of the value
        )
    except Exception:
        pass   # silently skip on non-Windows or older Windows

    app = DeathTrackerApp(root)

    # mainloop() hands control to tkinter's event loop.
    # It blocks here, processing clicks/keypresses/redraws
    # until the window is closed.
    root.mainloop()
