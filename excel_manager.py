# Excel
from openpyxl import Workbook, load_workbook

# File handling
import os

class ExcelManager:
    
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename, data_only=True)
        self.in_sheet = self.wb["input"]
        self.out_sheet = self.wb["output"]



    #------------------------#
    #    Records Deaths      #
    #------------------------#
    def record_session(self, game, deaths):
        """Append a death record"""

        target_col = None

        # Find game
        for col in range(1, self.in_sheet.max_column + 1):
            if self.in_sheet.cell(row=1, column=col).value == game:
                target_col = col
                break

        if target_col is None:
            raise ValueError(f"{game} not found in spreadsheet")
        
        # Find next empty row
        row = 2
        while self.in_sheet.cell(row=row, column=target_col).value is not None:
            row+=1
            
        # Record deaths
        self.in_sheet.cell(
            row=row,
            column=target_col
        ).value = deaths

        self.save()


    #------------------------#
    #   Grabs Needed PUs     #
    #------------------------#
    def get_pushups(self):
        return int(self.out_sheet["G2"].value or 0)


    #------------------------#
    #         Saves          #
    #------------------------#
    def save(self):
        """saves"""
        self.wb.save(self.filename)